# objective for this code: better use of functions.

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime, timedelta
import pyexcel
import os
from time import sleep


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)



def convert_xls(xls_file):
    """
    xls_file : xls file.
    
    ------------
        takes an XLS file. 
    ------------
        creates a converted XLSX file.
    
    """

    pyexcel.save_book_as(file_name=xls_file,
                         dest_file_name=xls_file+"x")


def delete_file(file):
    """
    file : any file.
    
    ----------
        takes a file.
    ----------
        deletes it. 
    
    """

    os.remove(file)


def scrape_table(worksheet, flag = 0):
    """
    Parameters
    ----------
    worksheet : openpyXL worksheet object.

    Returns
    -------
    List with the table contents.

    """
    # Row counter
    row_count = 0
    
    for row in worksheet:
        if not all([cell.value is None for cell in row]):
            row_count += 1

    # Loop through the rows and columns, appending the information to a dictionary.
    scraped_info = {}

    for r in range(1, row_count+1):

        for col in range(1, 5):
            char = get_column_letter(col)
            
            
            
            if ws["A" + str(r)].value == None:
                pass
            
            else:
                if char == "A":
                    scraped_info[ws[char + str(r)].value] = []
                else:
                    scraped_info[ws["A" +
                                    str(r)].value] += [ws[char + str(r)].value]

    return scraped_info


def write_table(client, worksheet, col_start=1):
    """
    Parameters
    ----------
    table : list with scraped data.
        
    worksheet : worksheet object to write new table onto.

    Returns
    -------
    None.
    """
    # write the tables.

    for i, key in enumerate(client):
        
        char = get_column_letter(col_start)
        ws[char+str(i+1)].value = key

        for j, value in enumerate(client[key]):
            char = get_column_letter(col_start+j+1)

            if char == "A":
                pass 
            else:
                ws[char+str(i+1)].value = client[key][j]


def date_conversion(date_string):
    """
    Parameters
    ----------
    date_string : String containing a date. example:
        "110223"  or  "11/02/21"  or  "blahblahbvlah 110423"

    Returns
    -------
    DateTime Object.
    """
    # strip the symbols in between the dates if existant.

    rawdate = date_string.split(" ")[1]
    y = int('20'+rawdate[4:])
    m = int(rawdate[0:2])
    d = int(rawdate[2:4])

    return datetime(y, m, d)



def date_conversion_rev(dt):
    """
    Parameters
    -------
    datetime: datetime object.

    Returns
    -------
    String containing a datestring in the following format: "YYMMDD"
    """
    
    return dt.strftime("%m%d%y")
    
    
def move_sheet(wb, from_loc=None, to_loc=None):
    sheets=wb._sheets

    # if no from_loc given, assume last sheet
    if from_loc is None:
        from_loc = len(sheets) - 1

    #if no to_loc given, assume first
    if to_loc is None:
        to_loc = 0

    sheet = sheets.pop(from_loc)
    sheets.insert(to_loc, sheet)
    sleep(0.5)
    
    
def get_file_names():
    """
    Returns
    -------
    All filenames of this python file's current directory, excluding itself. 
    
    """

    current_dir = os.getcwd()

    file_list = os.listdir(current_dir)

    file_list.remove("growth_report.py")
   

    return file_list


def sort_by_date(file_list):
    """
    Parameters
    ----------
    file_list : List of filenames with dates on them.

    Returns
    -------
    Sorted List By Date.
    """

    sorted_files = []
    sorted_dates = []
    
    
    
    for file in file_list:
        if file == "LAST_GROWTH_REPORT.xlsx":
            pass
        else:
            sorted_dates.append(date_conversion(file))

    sorted_dates.sort()
    
    
    # FROM THE CURRENT DATETIME, GET LAST WEEK'S DATETIME
    for file in file_list:
        if file == "LAST_GROWTH_REPORT.xlsx":
            
            #needs present
            last_week = get_last_week(sorted_dates[0])
            
            #LAST WEEK STRING SHOULD REFLECT THE OTHER FILES' FORMATS : Y M D
            last_week_string = date_conversion_rev(last_week)
            
            newfilename = "report "+last_week_string+" lastweek.xlsx"
            os.rename("LAST_GROWTH_REPORT.xlsx", newfilename )
            
            #remove old filename and add new filename
            file_list.remove("LAST_GROWTH_REPORT.xlsx")
            
            current_dir = os.getcwd()
            file_list2 = os.listdir(current_dir)
            
            for file in file_list2:
                if file == newfilename:
                    file_list.append(newfilename)
                    
                else:
                    pass            
            
            #add the datetime to sorted_dates
            sorted_dates.insert(0, last_week)
            
    
    
    for dt in sorted_dates:
        for file in file_list:
            if date_conversion(file) == dt:

                sorted_files.append(file)
        
    return sorted_files, sorted_dates


def get_last_week(datetime):
    """
    Parameters
    ----------
    datetime : datetime object. 

    Returns
    -------
    datetime_lastweek : datetime object containing last week's date in
    relation to the received datetime.
    """
    return (datetime - timedelta(days=7))



def get_week_id(dt):
    return dt.strftime("%V")

def get_week_str(dt):
    return dt.strftime("%m/%d/%y")



# CODE START

# convert all files to xls

file_list = get_file_names()

for file in file_list:
    if "xlsx" in file:
        pass
    else:
        convert_xls(file)
        delete_file(file)


# get list with converted files

file_list = get_file_names()
file_list.remove("rates-and-capacities.xlsx")


sorted_datetimes = sort_by_date(file_list)[0]


weeks_dt = sort_by_date(file_list)[1]


# scrape all the tables onto this list:

tables = []

for i, file in enumerate(sorted_datetimes):
    
    if i == 0:
        wb = load_workbook(file)
        ws = wb["Timeline"]
        
        tables.append(scrape_table(ws))
    else:
        wb = load_workbook(file)
        ws = wb.active
        
        tables.append(scrape_table(ws))


wb = load_workbook(sorted_datetimes[0])
ws = wb["Timeline"]

# OVERWRITE PAST REPORT'S TIMELINE
col = 1

f_tables = tables[:]
f_tables.pop(0)


for i, lst in enumerate(f_tables):
    
    write_table(lst, ws, col)
    col += 6



wb.save("result.xlsx")

# get the filename, open the worksheet, loop for the data. 

file_list = get_file_names()

for file in file_list:
    if file == "rates-and-capacities.xlsx":
        rates_table = file

wb = load_workbook(rates_table)
ws = wb.active

row_count = 0
 
for row in ws:
    if not all([cell.value is None for cell in row]):
        row_count += 1

FIXED_INFO = {}

for row in range(3, row_count+1):
    for column in range(1,5):
        char = get_column_letter(column)
        
        if column == 1:
            FIXED_INFO[(ws[char + str(row)].value)] = [ws[get_column_letter(column + 1) + str(row)].value, ws[get_column_letter(column + 2) + str(row)].value, ws[get_column_letter(column + 3) + str(row)].value]
            

wb = load_workbook("result.xlsx")
ws = wb["Timeline"]
            
#Change positions of the worksheets. Bring current one to top.
current_week_id = str(int(get_week_id(weeks_dt[1]))+1)

new_week_sheet = "Growth Report - Week " + current_week_id

wb.create_sheet(new_week_sheet)

move_sheet(wb, len(wb.sheetnames)-1, 1)

ws = wb[new_week_sheet]

titles = []

for i, key in enumerate(tables[5]):
    if i < 3:
        pass

    # take bellona out
    # canton and columbia dont exist in the originals?

    elif "999" in key:
        pass
    else:
        titles.append(key)


# CREATING BLOAT FREE TABLES LIST
org_tables = []

for dic in tables:

    org_tables.append(dic.copy())


for i, dic in enumerate(tables):

    for counter, row in enumerate(dic):
        
        if counter < 3:
            #remove item
            del org_tables[i][row]

        elif "999" in row:
            del org_tables[i][row]
            
        
data = []
            
for dic in org_tables:

    data.append(dic.copy())
    

for i, dic in enumerate(org_tables):
    
    for row in dic:
        
        #append values to the list 
        data[i][row] += FIXED_INFO[row]
        

# SEPARATE DATA IN "data" VARIABLE BY STATE

# FIRST CHALLENGE: HAVE CODE IDENTIFY AUTOMATICALLY HOW MANY DIFFERENT STATES ARE IN THE "data" variable:
    
data_states = []


for value in data[5].values():
    if value[5] not in data_states:
        data_states.append(value[5])


n_states = len(data_states)



# SECOND CHALLENGE: SEPARATE THE ACTUAL DATA
the_data = []
for l in range(0, n_states):
    the_data.append([])
    
    



for k, state in enumerate(data_states):
    # create variables using the state strings
    locals()[state] = []
    
    for i, dic in enumerate(data):
        check = 0
        
        for j, item in enumerate(dic.items()):
            
            # the filtering needs to happen HERE somehow. 
            # Only the schools matching the "state" variable 
            # should get through.
            
            if item[1][5] == state:
            
                if i == 0:
                    if check == 0:
                        the_data[k].append({item[0]:item[1]})
                        check = 1
                    else:
                        the_data[k][0][item[0]] = item[1]
                    
                    
                if i == 1:
                    if check == 0:
                        the_data[k].append({item[0]:item[1]})
                        check = 1
                    else:
                        the_data[k][1][item[0]] = item[1]
                    
                    
                if i == 2:
                    if check == 0:
                        the_data[k].append({item[0]:item[1]})
                        check = 1
                    else:
                        the_data[k][2][item[0]] = item[1]
                    
                    
                if i == 3:
                    if check == 0:
                        the_data[k].append({item[0]:item[1]})
                        check = 1
                    else:
                        the_data[k][3][item[0]] = item[1]
                    
                    
                if i == 4:
                    if check == 0:
                        the_data[k].append({item[0]:item[1]})
                        check = 1
                    else:
                        the_data[k][4][item[0]] = item[1]
                    
                    
                if i == 5:
                    if check == 0:
                        the_data[k].append({item[0]:item[1]})
                        check = 1
                    else:
                        the_data[k][5][item[0]] = item[1]


# GET RE_ORGANIZED TITLES                        
state_keys = []
                    
for j, state in enumerate(the_data):
    
    for i, key in enumerate(state[5].keys()):
        if i == 0:
            state_keys.append([])
        
        state_keys[j].append(key)


# DRAW THE TITLES ON THE TABLES

counter = 1
for j, state in enumerate(state_keys):
    
    ws["A" + str(counter)].value = data_states[j]
    counter += 1
    
    for i in range(0, len(state)):
    
        ws["A" + str(counter)].value = state[i]
        counter += 1



counter = 1
for j, state in enumerate(state_keys):
    
    ws["E" + str(counter)].value = data_states[j]
    counter += 1
    
    for i in range(0, len(state)):
    
        ws["E" + str(counter)].value = state[i]
        counter += 1
    



table_len = len(titles)      
  
counter = 1
for j, state in enumerate(state_keys):
    
    ws["E" + str(counter + table_len + n_states + 4)].value = data_states[j]
    counter += 1
    
    for i in range(0, len(state)):
    
        ws["E" + str(counter + table_len + n_states + 4)].value = state[i]
        counter += 1
    
counter = 1
for j, state in enumerate(state_keys):
    
    ws["A" + str(counter + table_len + n_states + 4)].value = data_states[j]
    counter += 1
    
    for i in range(0, len(state)):
    
        ws["A" + str(counter + table_len + n_states + 4)].value = state[i]
        counter += 1


# DATA TABLE    

counter = 1

for j, data_ in enumerate(the_data):
    if j == 0:
        id_a = 0
    else:
        id_a += len(the_data[j-1][5])+1
   
    
    for a, dic in enumerate(data_):
        b = 0  
        for c, key in enumerate(dic):
            
        
            for col in range(6, 11):
                char = get_column_letter(col)
                
                if char == "F" and a == 1:
                    ws[char + str(b+2+ id_a)].value = int(dic[key][0]/dic[key][3])
                    ws[char + str(b+ id_a+table_len+ n_states + 6)].value = str(int((dic[key][0]/dic[key][3]) * 100 /dic[key][4]))+"%"
                    ws[char + str(b+ id_a+table_len+ n_states + 6)].alignment = Alignment(horizontal = "right")
                
                if char == "G" and a == 2:
                    ws[char + str(b+2+ id_a)].value = int(dic[key][0]/dic[key][3])
                    ws[char + str(b+ id_a+table_len+ n_states + 6)].value = str(int((dic[key][0]/dic[key][3]) * 100 /dic[key][4]))+"%"
                    ws[char + str(b+ id_a+table_len+ n_states + 6)].alignment = Alignment(horizontal = "right")
                    
                if char == "H" and a == 3:
                    ws[char + str(b+2+ id_a)].value = int(dic[key][0]/dic[key][3])
                    ws[char + str(b+ id_a+table_len+ n_states + 6)].value = str(int((dic[key][0]/dic[key][3]) * 100 /dic[key][4]))+"%"
                    ws[char + str(b+ id_a+table_len+ n_states + 6)].alignment = Alignment(horizontal = "right")
                    
                if char == "I" and a == 4:
                    ws[char + str(b+2+ id_a)].value = int(dic[key][0]/dic[key][3])
                    ws[char + str(b+ id_a+table_len+ n_states + 6)].value = str(int((dic[key][0]/dic[key][3]) * 100 /dic[key][4]))+"%"
                    ws[char + str(b+ id_a+table_len+ n_states + 6)].alignment = Alignment(horizontal = "right")
                    
                if char == "J" and a == 5:
                    ws[char + str(b+2+ id_a)].value = int(dic[key][0]/dic[key][3])
                    ws[char + str(b+ id_a+table_len+ n_states + 6)].value = str(int((dic[key][0]/dic[key][3]) * 100 /dic[key][4]))+"%"
                    ws[char + str(b+ id_a+table_len+ n_states + 6)].alignment = Alignment(horizontal = "right")
            b += 1
    #??? += len(data_[5])+1    


for j, data_ in enumerate(the_data):
    
    if j == 0:
        id_a = 0
    else:
        id_a += len(the_data[j-1][5])+1
        
    for a, dic in enumerate(data_):
        
        b = 0
    
        for c, key in enumerate(dic):
            
            for col in range(2, 4):
                char = get_column_letter(col)
                
                if a == 1 and col == 2:
                    ws[char + str(b+2+ id_a)].value = int(dic[key][0]/dic[key][3])
                    
                    
                elif a == 1 and col == 3:
                    try:
                        ws[char + str(b+2+ id_a)].value = int(dic[key][0]/dic[key][3]) - int(data[a-1][key][0]/dic[key][3])
                    except:
                        pass
                    
                elif a == 5 and char == "B":
                    ws[char + str(b+table_len+ id_a + n_states + 6)].value = dic[key][4]
            b += 1
    #??? += len(data_[5])+1      
        





 ###STYLE###

for column in range(1, 25):
    for row in range(1, 60):
        char = get_column_letter(column)
        ws[char+str(row)].font = Font(name = "Calibri Light")
        


ws.move_range("A1:J" + str(((table_len+n_states)*2)+4), rows=3, cols=3)         #36 needs to be a variable that accounts for the number of schools

ws.row_dimensions[1].height = 24
#ws.row_dimensions[table_len + 4].height = 24

ws["D1"].font = Font(size = 19)
ws["D1"].alignment = Alignment(horizontal='center')

ws["D1"].fill = PatternFill(fill_type='solid',
                            start_color='99ffcc',
                            end_color='99ffcc')

ws["H1"].font = Font(size=19)
ws["H1"].alignment = Alignment(horizontal='center')
ws["H1"].fill = PatternFill(fill_type='solid',
                            start_color='99ffcc',
                            end_color='99ffcc')

ws["D"+str(table_len+n_states + 7)].font = Font(size = 19)
ws["D"+str(table_len+n_states + 7)].alignment = Alignment(horizontal='center')
ws["d"+str(table_len+n_states + 7)].fill = PatternFill(fill_type='solid',
                            start_color='ccffcc',
                            end_color='ccffcc')

ws["H"+str(table_len+n_states + 7)].font = Font(size = 19)
ws["H"+str(table_len+n_states + 7)].alignment = Alignment(horizontal='center')
ws["H"+str(table_len+n_states + 7)].fill = PatternFill(fill_type='solid',
                            start_color='ccffcc',
                            end_color='ccffcc')


ws.column_dimensions["D"].width = 35
ws.column_dimensions["F"].width = 15
ws.column_dimensions["H"].width = 35

ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 27
ws.column_dimensions["C"].width = 3

ws.row_dimensions[2].height = 12.5
ws["E2"].font = Font(size = 8)
ws["I2"].font = Font(size = 8)

ws.merge_cells("D1:F1")
ws.merge_cells("H1:M1")
ws.merge_cells("D"+str(table_len+n_states + 7)+":E"+str(table_len+n_states + 7))
ws.merge_cells("H"+str(table_len+n_states + 7)+":M"+str(table_len+n_states + 7))

ws["D1"].value = "FTE Children"

ws["H1"].value = "FTE Children BD Projections"
ws["D"+str(table_len+n_states + 7)].value = "School Max Capacity"
ws["H"+str(table_len+n_states + 7)].value = "School Occupancy"


ws["E2"].value = "Current"
ws["D3"].value = "School ID"
ws["E3"].value = get_week_str(weeks_dt[1])
ws["F3"].value = "Growth from LW"

ws["I2"].value = "Current"
ws["H3"].value = "School ID"
ws["I3"].value = get_week_str(weeks_dt[1])
ws["J3"].value = get_week_str(weeks_dt[2])
ws["K3"].value = get_week_str(weeks_dt[3])
ws["L3"].value = get_week_str(weeks_dt[4])
ws["M3"].value = get_week_str(weeks_dt[5])

ws.merge_cells("B1:B2")
ws["B1"].value = "Growth Report"
ws["B1"].font = Font(size = 19)
ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
ws["B1"].fill = PatternFill(fill_type='solid',
                            start_color='ccffcc',
                            end_color='ccffcc')


ws["B3"].value = "Week "+ current_week_id +" of 52"
ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")


####BORDERS####
set_border(ws, "B1:B3")
set_border(ws, 'D1:F'+ str(table_len+n_states+3))        #ALL FOUR OF THESE ALSO NEED THE VARIABLES WITH THE AMOUNT OF SCHOOLS
set_border(ws, 'D'+str(table_len +n_states+ 7)+':E'+ str((table_len+n_states+3)*2+1)) 
set_border(ws, 'H1:M'+ str(table_len+n_states+3)) 
set_border(ws, 'H'+str(table_len +n_states+ 7)+':M' + str((table_len+n_states+3)*2+1)) 


wb.save("result.xlsx")

