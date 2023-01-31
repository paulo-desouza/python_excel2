# objective for this code: better use of functions.
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
import pyexcel
import os


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)



def convert_xls(xls_file):
    """
    xls_file : xls file.
    
    ------------
        takes an XLS file. 
    ------------
        creates a converted XLSX file.
    
    """

    pyexcel.save_book_as(file_name=xls_file,
                         dest_file_name=xls_file+"x")


def delete_file(file):
    """
    file : any file.
    
    ----------
        takes a file.
    ----------
        deletes it. 
    
    """

    os.remove(file)


def scrape_table(worksheet):
    """
    Parameters
    ----------
    worksheet : openpyXL worksheet object.

    Returns
    -------
    List with the table contents.

    """
    # Row and column counter
    row_count = 0

    for row in worksheet:
        if not all([cell.value is None for cell in row]):
            row_count += 1

    # Loop through the rows and columns, appending the information to a list.
    scraped_info = {}

    for r in range(1, row_count+1):

        for col in range(1, 5):

            char = get_column_letter(col)

            if char == "A":
                scraped_info[worksheet[char + str(r)].value] = []
            else:
                scraped_info[worksheet["A" +
                                str(r)].value] += [worksheet[char + str(r)].value]

    return scraped_info


def write_table(client, worksheet, col_start=1):
    """
    Parameters
    ----------
    table : list with scraped data.
        
    worksheet : worksheet object to write new table onto.

    Returns
    -------
    None.
    """
    # write the tables.

    for i, key in enumerate(client):
        char = get_column_letter(col_start)
        ws[char+str(i+1)].value = key

        for j, value in enumerate(client[key]):
            char = get_column_letter(col_start+j+1)

            if char == "A":
                pass
            else:
                ws[char+str(i+1)].value = client[key][j]


def date_conversion(date_string):
    """
    Parameters
    ----------
    date_string : String containing a date. example:
        "110223"  or  "11/02/21"  or  "blahblahbvlah 110423"

    Returns
    -------
    DateTime Object.
    """
    # strip the symbols in between the dates if existant.

    rawdate = date_string.split(" ")[1]
    y = int('20'+rawdate[4:])
    m = int(rawdate[0:2])
    d = int(rawdate[2:4])

    return datetime(y, m, d)


def get_file_names():
    """
    Returns
    -------
    All filenames of this python file's current directory, excluding itself. 
    
    """

    current_dir = os.getcwd()

    file_list = os.listdir(current_dir)

    file_list.remove("first_of_the_year.py")

    return file_list


def get_last_week_data(file_list):
    
    for file in file_list:
        
        if file == "last_yearly_report.xlsx":
            
            wb = load_workbook(file)
            ws = wb["Timeline"]
            
            table_info = scrape_table(ws)
            
            return table_info

def sort_by_date(file_list):
    """
    Parameters
    ----------
    file_list : List of filenames with dates on them.

    Returns
    -------
    Sorted List By Date.
    """

    sorted_files = []
    sorted_dates = []

    for file in file_list:
        if file == "last_yearly_report.xlsx":
            file_list.remove(file)
        else:    
            sorted_dates.append(date_conversion(file))

    sorted_dates.sort()

    for dt in sorted_dates:
        for file in file_list:

            if date_conversion(file) == dt:

                sorted_files.append(file)

    return sorted_files, sorted_dates


def get_week_id(dt):
    return dt.strftime("%V")

def get_week_str(dt):
    return dt.strftime("%m/%d/%y")



# CODE START

# convert all files to xls

file_list = get_file_names()


#get_last_weeks table
last_week_table = get_last_week_data(file_list)

for file in file_list:
    if "xlsx" in file:
        pass
    else:
        convert_xls(file)
        delete_file(file)

# get list with converted files

file_list = get_file_names()
file_list.remove("rates-and-capacities.xlsx")


sorted_datetimes = sort_by_date(file_list)[0]

weeks_dt = sort_by_date(file_list)[1]


# scrape all the tables onto this list:
tables = []

for i, file in enumerate(sorted_datetimes):
    wb = load_workbook(file)
    ws = wb.active

    tables.append(scrape_table(ws))
    
    
# FIXED INFO GATHERING
file_list = get_file_names()

for file in file_list:
    if file == "rates-and-capacities.xlsx":
        rates_table = file
        
        
wb = load_workbook(rates_table)
ws = wb.active

row_count = 0
 
for row in ws:
    if not all([cell.value is None for cell in row]):
        row_count += 1

FIXED_INFO = {}

for row in range(3, row_count+1):
    for column in range(1,4):
        char = get_column_letter(column)
        
        if column == 1:
            FIXED_INFO[(ws[char + str(row)].value)] = [ws[get_column_letter(column + 1) + str(row)].value, ws[get_column_letter(column + 2) + str(row)].value]


wb = Workbook()
ws = wb.active
ws.title = "Timeline"


col = 1
for list in tables:
    write_table(list, ws, col)
    col += 6


            

wb.create_sheet("Growth Report")
ws = wb["Growth Report"]


titles = []

for i, key in enumerate(tables[4]):
    if i < 3:
        pass

    # take bellona out
    # canton and columbia dont exist in the originals?

    elif "999" in key:
        pass
    else:
        titles.append(key)


# CREATING BLOAT FREE TABLES LIST
org_tables = []
org_LW = []

for dic in tables:

    org_tables.append(dic.copy())
    
org_LW.append(last_week_table.copy())


for i, dic in enumerate(tables):

    for counter, row in enumerate(dic):
        
        if counter < 3:
            #remove item
            del org_tables[i][row]

        elif "999" in row:
            del org_tables[i][row]
            


for counter, row in enumerate(last_week_table):
    if row == None:
        del org_LW[0][row]
    else:
        if counter < 3:
            #remove item
            del org_LW[0][row]
    
        elif "999" in row:
            del org_LW[0][row]            
    
        
data = []
LW_data = []


for dic in org_tables:

    data.append(dic.copy())
    

for i, dic in enumerate(org_tables):
    
    for row in dic:
        
        #append values to the list 
        data[i][row] += FIXED_INFO[row]
        
        
for dic in org_LW:

    LW_data.append(dic.copy())
    

for i, dic in enumerate(org_LW):
    
    for row in dic:
        
        #append values to the list 
        LW_data[i][row] += FIXED_INFO[row]

#TITLES

for i in range(0, len(titles)):
    
    if i==0:
        ws["A" + str(i+1)].value = titles[i]
    else:
        ws["A" + str(i+1)].value = titles[i]

for i in range(0, len(titles)):

    ws["E" + str(i+1)].value = titles[i]

table_len = len(titles)

for a, dic in enumerate(data):

    for b, key in enumerate(dic):
        
        for col in range(2, 4):
            char = get_column_letter(col)
            
            if a == 0 and col == 2:
                ws[char + str(b+1)].value = int(dic[key][0]/dic[key][3])
                
                
            elif a == 0 and col == 3:
                
                ws[char + str(b+1)].value = int(dic[key][0]/dic[key][3]) - int(last_week_table[key][0]/last_week_table[key][3])
                
            elif a == 4 and char == "B":
                ws[char + str(b + table_len + 4)].value = dic[key][4]

for a, dic in enumerate(data):

    for b, key in enumerate(dic):
        
        for col in range(6, 11):
            char = get_column_letter(col)
            
            if char == "F" and a == 0:
                ws[char + str(b+1)].value = int(dic[key][0]/dic[key][3])
                ws[char + str(b+ table_len + 4)].value = str(int((dic[key][0]/dic[key][3]) * 100 /dic[key][4]))+"%"
                ws[char + str(b+ table_len + 4)].alignment = Alignment(horizontal = "right")
            
            if char == "G" and a == 1:
                ws[char + str(b+1)].value = int(dic[key][0]/dic[key][3])
                ws[char + str(b+ table_len + 4)].value = str(int((dic[key][0]/dic[key][3]) * 100 /dic[key][4]))+"%"
                ws[char + str(b+ table_len + 4)].alignment = Alignment(horizontal = "right")
                
            if char == "H" and a == 2:
                ws[char + str(b+1)].value = int(dic[key][0]/dic[key][3])
                ws[char + str(b+ table_len + 4)].value = str(int((dic[key][0]/dic[key][3]) * 100 /dic[key][4]))+"%"
                ws[char + str(b+ table_len + 4)].alignment = Alignment(horizontal = "right")
                
            if char == "I" and a == 3:
                ws[char + str(b+1)].value = int(dic[key][0]/dic[key][3])
                ws[char + str(b+ table_len + 4)].value = str(int((dic[key][0]/dic[key][3]) * 100 /dic[key][4]))+"%"
                ws[char + str(b+ table_len + 4)].alignment = Alignment(horizontal = "right")
                
            if char == "J" and a == 4:
                ws[char + str(b+1)].value = int(dic[key][0]/dic[key][3])
                ws[char + str(b+ table_len + 4)].value = str(int((dic[key][0]/dic[key][3]) * 100 /dic[key][4]))+"%"
                ws[char + str(b+ table_len + 4)].alignment = Alignment(horizontal = "right")
        
        
        
for i in range(0, len(titles)):

    ws["A" + str(i+ table_len + 4)].value = titles[i]

        
for i in range(0, len(titles)):

    ws["E" + str(i+ table_len + 4)].value = titles[i]


                ###STYLE###

for column in range(1, 25):
    for row in range(1, 60):
        char = get_column_letter(column)
        ws[char+str(row)].font = Font(name = "Calibri Light")
        

table_len = len(titles)


ws.move_range("A1:J" + str((table_len*2)+3), rows=3, cols=3)

ws.row_dimensions[1].height = 24
#ws.row_dimensions[table_len + 4].height = 24

ws["D1"].font = Font(size = 19)
ws["D1"].alignment = Alignment(horizontal='center')

ws["D1"].fill = PatternFill(fill_type='solid',
                            start_color='99ffcc',
                            end_color='99ffcc')

ws["H1"].font = Font(size=19)
ws["H1"].alignment = Alignment(horizontal='center')
ws["H1"].fill = PatternFill(fill_type='solid',
                            start_color='99ffcc',
                            end_color='99ffcc')

ws["D"+ str(table_len + 6)].font = Font(size = 19)
ws["D"+ str(table_len + 6)].alignment = Alignment(horizontal='center')
ws["D"+ str(table_len + 6)].fill = PatternFill(fill_type='solid',
                            start_color='ccffcc',
                            end_color='ccffcc')

ws["H"+ str(table_len + 6)].font = Font(size = 19)
ws["H"+ str(table_len + 6)].alignment = Alignment(horizontal='center')
ws["H"+ str(table_len + 6)].fill = PatternFill(fill_type='solid',
                            start_color='ccffcc',
                            end_color='ccffcc')


ws.column_dimensions["D"].width = 35
ws.column_dimensions["F"].width = 15
ws.column_dimensions["H"].width = 35

ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 27
ws.column_dimensions["C"].width = 3

ws.row_dimensions[2].height = 12.5
ws["E2"].font = Font(size = 8)
ws["I2"].font = Font(size = 8)

ws.merge_cells("D1:F1")
ws.merge_cells("H1:M1")
ws.merge_cells("D"+ str(table_len + 6)+":E"+ str(table_len + 6))
ws.merge_cells("H"+ str(table_len + 6)+":M"+ str(table_len + 6))

ws["D1"].value = "FTE Children"

ws["H1"].value = "FTE Children BD Projections"
ws["D"+ str(table_len + 6)].value = "School Max Capacity"
ws["H"+ str(table_len + 6)].value = "School Occupancy"


ws["E2"].value = "Current"
ws["D3"].value = "School ID"
ws["E3"].value = get_week_str(weeks_dt[0])
ws["F3"].value = "Growth from LW"

ws["I2"].value = "Current"
ws["H3"].value = "School ID"
ws["I3"].value = get_week_str(weeks_dt[0])
ws["J3"].value = get_week_str(weeks_dt[1])
ws["K3"].value = get_week_str(weeks_dt[2])
ws["L3"].value = get_week_str(weeks_dt[3])
ws["M3"].value = get_week_str(weeks_dt[4])

ws.merge_cells("B1:B2")
ws["B1"].value = "Growth Report"
ws["B1"].font = Font(size = 19)
ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
ws["B1"].fill = PatternFill(fill_type='solid',
                            start_color='ccffcc',
                            end_color='ccffcc')


ws["B3"].value = "Week 1 of 52"
ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")


####BORDERS####
set_border(ws, "B1:B3")
set_border(ws, 'D1:F'+ str(table_len+3))        #ALL FOUR OF THESE ALSO NEED THE VARIABLES WITH THE AMOUNT OF SCHOOLS
set_border(ws, 'D'+ str(table_len + 6)+':E'+ str((table_len+3)*2)) 
set_border(ws, 'H1:M'+ str(table_len+3)) 
set_border(ws, 'H'+ str(table_len + 6)+':M' + str((table_len+3)*2)) 


wb.save("result.xlsx")




























